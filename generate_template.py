"""
Скрипт для создания Excel-шаблона учёта заказов Ozon
с условным форматированием, выпадающими списками и готовой структурой.

Как использовать:
1. Установите Python (https://www.python.org/downloads/)
2. Откройте терминал (Win+R → cmd)
3. Выполните: pip install openpyxl
4. Выполните: python generate_template.py
5. Откройте созданный файл "Ozon_Заказы_шаблон.xlsx"
"""

import os
import sys

# Windows-консоль (cp1251) не может вывести эмодзи — переключаем stdout на UTF-8.
if sys.platform == 'win32' and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# --- Конфигурация ---
OUTPUT = "Ozon_Заказы_шаблон.xlsx"
SHEET_NAME = "Заказы"

# Цвета
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Столбцы (ширина, название) — единый словарь заголовков B1 (10 колонок).
# K — служебная колонка со справкой (в автофильтр не входит).
COLUMNS = [
    ("Дата заказа", 14),
    ("№ Заказа", 18),
    ("Статус доставки", 24),
    ("Товары", 50),
    ("Кол-во", 10),
    ("Сумма", 12),
    ("Статус оплаты", 22),
    ("Пункт выдачи", 28),
    ("Дата доставки", 16),
    ("Фото", 40),
]

# Список статусов для выпадающих списков
DELIVERY_STATUSES = [
    "📦 Обрабатывается",
    "📦 Собирается",
    "🚚 Передан в доставку",
    "🚚 В пути",
    "📦 Готов к выдаче",
    "✅ Доставлен",
    "❌ Отменён",
]

PAYMENT_STATUSES = [
    "✅ Оплачен",
    "❌ Не оплачен",
    "⏳ При получении",
    "🟡 Частично оплачен",
    "🔄 Возврат",
    "⏳ Ожидает оплаты",
]


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # --- Заголовки ---
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Служебная колонка K со справкой (ширина ~40)
    ws.column_dimensions['K'].width = 40

    # --- Фиксация шапки ---
    ws.freeze_panes = "A2"

    # --- Автофильтр A1:J (K — служебная, не фильтруется) ---
    ws.auto_filter.ref = f"A1:J1"

    # --- Строка-подсказка (скрытая) ---
    ws.cell(row=2, column=1, value="Вставьте данные сюда (Ctrl+V)")
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")

    # --- Выпадающие списки (Data Validation) для 1000 строк ---
    # Статус доставки (столбец C)
    dv_delivery = DataValidation(
        type="list",
        formula1=f'"{",".join(DELIVERY_STATUSES)}"',
        allow_blank=True
    )
    dv_delivery.error = "Выберите статус из списка"
    dv_delivery.errorTitle = "Неверное значение"
    dv_delivery.prompt = "Выберите статус доставки"
    dv_delivery.promptTitle = "Статус доставки"
    ws.add_data_validation(dv_delivery)
    dv_delivery.add(f"C2:C1001")

    # Статус оплаты (столбец G)
    dv_payment = DataValidation(
        type="list",
        formula1=f'"{",".join(PAYMENT_STATUSES)}"',
        allow_blank=True
    )
    dv_payment.error = "Выберите статус из списка"
    dv_payment.errorTitle = "Неверное значение"
    dv_payment.prompt = "Выберите статус оплаты"
    dv_payment.promptTitle = "Статус оплаты"
    ws.add_data_validation(dv_payment)
    dv_payment.add(f"G2:G1001")

    # --- Условное форматирование ---
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")

    # Сумма <= 0 или пусто — светло-жёлтый (не оплачено)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C5700")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    # Столбец F (Сумма): если число > 0 — зелёный фон
    ws.conditional_formatting.add(
        f"F2:F1001",
        CellIsRule(
            operator='greaterThan',
            formula=['0'],
            fill=green_fill,
            font=green_font
        )
    )

    # Столбец F (Сумма): если 0 — жёлтый
    ws.conditional_formatting.add(
        f"F2:F1001",
        CellIsRule(
            operator='lessThanOrEqual',
            formula=['0'],
            fill=yellow_fill,
            font=yellow_font
        )
    )

    # Столбец F (Сумма): пустые ячейки — жёлтый (CellIsRule не покрывает пустые,
    # добавляем FormulaRule с ISBLANK — B3 п.4)
    ws.conditional_formatting.add(
        f"F2:F1001",
        FormulaRule(
            formula=['ISBLANK(F2)'],
            fill=yellow_fill,
            font=yellow_font
        )
    )

    # Столбец G (Статус оплаты): содержит "✅ Оплачен" — зелёный
    ws.conditional_formatting.add(
        f"G2:G1001",
        CellIsRule(
            operator='equal',
            formula=['"✅ Оплачен"'],
            fill=green_fill,
            font=green_font
        )
    )

    # Столбец G (Статус оплаты): содержит "❌ Не оплачен" — красный
    ws.conditional_formatting.add(
        f"G2:G1001",
        CellIsRule(
            operator='equal',
            formula=['"❌ Не оплачен"'],
            fill=red_fill,
            font=red_font
        )
    )

    # Столбец C (Статус доставки): содержит "✅ Доставлен" — зелёный
    ws.conditional_formatting.add(
        f"C2:C1001",
        CellIsRule(
            operator='equal',
            formula=['"✅ Доставлен"'],
            fill=green_fill,
            font=green_font
        )
    )

    # --- Справка (столбец K) ---
    ws.cell(row=1, column=11, value="🔥 Справка по формату")
    ws.cell(row=1, column=11).font = Font(bold=True, size=11)

    help_text = [
        "Скопируйте данные через кнопку на сайте Ozon",
        "",
        "📌 Структура данных (v9 — 10 колонок):",
        "Каждый товар на отдельной строке со своей ценой.",
        "Номер заказа — только в первой строке.",
        "",
        "📌 Столбцы:",
        "A — Дата заказа (заполняется автоматически, из статуса или fallback)",
        "B — № Заказа (только в первой строке)",
        "C — Статус доставки (выпадающий список)",
        "D — Товары (каждый товар на своей строке)",
        "E — Кол-во (количество единиц товара, из корзины)",
        "F — Сумма (индивидуальная цена товара)",
        "G — Статус оплаты (выпадающий список, в каждой строке)",
        "H — Пункт выдачи (в каждой строке)",
        "I — Дата доставки (ожидаемая, из shipment widget)",
        "J — Фото (ссылка на изображение товара с ozon.ru)",
        "",
        "💡 Советы:",
        "- Используйте автофильтр (▲ в шапке)",
        "- Общая сумма без отменённых: =СУММЕСЛИ(C:C;\"<>❌ Отменён\";F:F)",
        "- Количество заказов: =СЧЁТЗ(B:B)",
        "- Сводка по статусам: =СЧЁТЕСЛИ(C:C,\"✅ Доставлен\")",
        "",
        "⚠️ ВАЖНО (v9.15):",
        "Старые шаблоны (11 колонок, «Примечание» в K) НЕ совместимы.",
        "Используйте новый шаблон, созданный этим скриптом.",
    ]
    for i, line in enumerate(help_text, 2):
        ws.cell(row=i, column=11, value=line)
        ws.cell(row=i, column=11).font = Font(size=10, color="333333")

    # --- Сохранение ---
    wb.save(OUTPUT)
    print(f"✅ Шаблон создан: {os.path.abspath(OUTPUT)}")
    print()
    print("📋 Дальнейшие действия:")
    print("1. Откройте файл в Excel или Yandex Tables")
    print("2. Скопируйте заказы с Ozon (кнопка на странице)")
    print("3. Вставьте в таблицу (Ctrl+V)")
    print("4. Статусы можно менять через выпадающие списки")
    print("5. Строки подсвечиваются цветом в зависимости от статуса")


if __name__ == "__main__":
    create_workbook()
